import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, numbers
from openpyxl.utils import get_column_letter


def count_hours_per_day(csv):
    table = pd.read_csv(csv, sep=';', usecols=['Сотрудник', 'Время, часы'])
    table['Время, часы'] = table['Время, часы'].str.replace(',', '.').astype(float)
    table = table.groupby('Сотрудник')['Время, часы'].sum().reset_index()
    table = table.sort_values(by='Сотрудник', key=lambda x: x.str.split().str[-1])
    result = table.reset_index(drop=True)
    return result


def put_data(data, date):
    letter = get_column_letter(date + 3)
    workbook = load_workbook('example.xlsx')

    if data['Сотрудник'] not in workbook.sheetnames:
        create_sheet(workbook, data)
        sheet = workbook[f'{data["Сотрудник"]}']
        sheet[f'{letter}4'] = data['Время, часы']
        workbook.save('example.xlsx')
    else:
        sheet = workbook[f'{data["Сотрудник"]}']
        sheet[f'{letter}4'] = data['Время, часы']
        workbook.save('example.xlsx')


def create_sheet(workbook, data):

    sheet = workbook.create_sheet()
    sheet.title = data['Сотрудник']

    # Columns Width
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 32
    for column_index in range(4, 34):
        column_letter = get_column_letter(column_index)
        column_dimensions = sheet.column_dimensions[column_letter]
        column_dimensions.width = 3

    # Rows Height
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 28
    sheet.row_dimensions[3].height = 10
    sheet.row_dimensions[4].height = 44

    # Styles
    text_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
    font1 = Font(name='Verdana', size=6, bold=False, italic=False, color='000000')
    font2 = Font(name='Verdana', size=8, bold=False, italic=False, color='000000')

    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    cell_range = sheet['A1:AG4']
    for row in cell_range:
        for cell in row:
            cell.alignment = text_style
            cell.font = font1
            cell.border = border

    cell_range = sheet['A4:C4']
    for row in cell_range:
        for cell in row:
            cell.font = font2

    # Data
    sheet['A1'] = 'Номер\n по\n порядку'
    sheet['B1'] = 'Фамилия, инициалы,\n должность\n (специальность,\n профессия)'
    sheet['C1'] = 'Табельный\n номер'
    sheet['D1'] = 'Отметки о явках и неявках на работу по числам месяца'
    sheet['A3'] = 1
    sheet['B3'] = 2
    sheet['C3'] = 3
    sheet['D3'] = 4
    sheet['A4'] = data.name + 1
    sheet['B4'] = data['Сотрудник']

    for column_index in range(4, 34):
        column_letter = get_column_letter(column_index)
        cell = sheet[f'{column_letter}{2}']
        cell.value = int(column_index - 3)

    # Merging
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:B2')
    sheet.merge_cells('C1:C2')
    sheet.merge_cells('D1:AG1')
    sheet.merge_cells('D3:AG3')


file = '02.07.2023.csv'
day = int(file[:2])
df = count_hours_per_day(file)
for i in range(len(df)):
    put_data(df.loc[i], day)
